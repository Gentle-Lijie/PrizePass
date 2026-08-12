import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook


MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_ROWS = 10_000
PRIZE_HEADERS = ["name", "image", "real_value", "redeem_value", "stock", "description"]
INTEGER_RE = re.compile(r"^(0|[1-9]\d*)$")
MONEY_RE = re.compile(r"^(0|[1-9]\d*)(?:\.(\d{1,2}))?$")


@dataclass(slots=True)
class TableData:
    headers: list[str]
    rows: list[list[Any]]


def read_table(filename: str, content: bytes) -> TableData:
    if len(content) > MAX_FILE_SIZE:
        raise ValueError("文件不能超过 5 MB")
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV 必须使用 UTF-8 编码") from exc
        matrix = list(csv.reader(io.StringIO(decoded)))
    elif suffix == "xlsx":
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            matrix = [list(row) for row in workbook.worksheets[0].iter_rows(values_only=True)]
            workbook.close()
        except Exception as exc:
            raise ValueError("无法读取 XLSX 文件") from exc
    else:
        raise ValueError("只支持 CSV 或 XLSX 文件")
    if not matrix:
        raise ValueError("表格不能为空")
    headers = [str(cell).strip() if cell is not None else "" for cell in matrix[0]]
    rows = [row for row in matrix[1:] if any(cell not in (None, "") for cell in row)]
    if len(rows) > MAX_ROWS:
        raise ValueError("表格最多包含 10,000 行，请拆分后重试")
    return TableData(headers=headers, rows=rows)


def parse_nonnegative_integer(value: Any, label: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label}必须是非负整数")
    text = str(value).strip() if value is not None else ""
    if isinstance(value, int) and value >= 0:
        return value
    if not INTEGER_RE.fullmatch(text):
        raise ValueError(f"{label}必须是非负整数")
    return int(text)


def parse_positive_integer(value: Any, label: str) -> int:
    parsed = parse_nonnegative_integer(value, label)
    if parsed <= 0:
        raise ValueError(f"{label}必须是大于 0 的整数")
    return parsed


def parse_money_to_cents(value: Any) -> int:
    text = str(value).strip() if value is not None else ""
    if not MONEY_RE.fullmatch(text):
        raise ValueError("真实价值必须是最多两位小数的非负金额")
    try:
        cents = int(Decimal(text) * 100)
    except InvalidOperation as exc:
        raise ValueError("真实价值格式错误") from exc
    if cents > 4_294_967_295:
        raise ValueError("真实价值过大")
    return cents


def safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def export_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows([[safe_cell(value) for value in row] for row in rows])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def export_xlsx(headers: list[str], rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(headers)
    for row in rows:
        sheet.append([safe_cell(value) for value in row])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def template_bytes(headers: list[str], format_name: str) -> tuple[bytes, str]:
    rows: list[list[Any]] = []
    if format_name == "csv":
        return export_csv(headers, rows), "text/csv; charset=utf-8"
    if format_name == "xlsx":
        return export_xlsx(headers, rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    raise ValueError("format 必须是 csv 或 xlsx")
