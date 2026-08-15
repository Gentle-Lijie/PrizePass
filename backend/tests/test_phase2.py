import csv
import io

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.main import app
from app.models import Event, Prize
from tests.conftest import test_engine


client = TestClient(app)
ADMIN = {"X-Admin-Password": "prizepass-dev-admin"}


def event_payload(status: str = "draft") -> dict:
    return {
        "name": "秋季编程赛",
        "description": "面向校内选手",
        "redemption_deadline": "2026-12-31T12:00:00Z",
        "pickup_location": "科技楼一层服务台",
        "pickup_instructions": "工作日 10:00-17:00 凭兑换单领取",
        "status": status,
    }


def prize_payload() -> dict:
    return {
        "name": "保温杯",
        "image": "https://example.com/cup.jpg",
        "real_value": 19900,
        "redeem_value": 150,
        "stock": 20,
        "description": "黑色保温杯",
    }


def create_event() -> int:
    response = client.post("/api/admin/events", headers=ADMIN, json=event_payload())
    assert response.status_code == 201, response.text
    return response.json()["id"]


def test_event_and_six_field_prize_crud() -> None:
    event_id = create_event()
    active = event_payload("active")
    response = client.put(f"/api/admin/events/{event_id}", headers=ADMIN, json=active)
    assert response.status_code == 200
    assert response.json()["status"] == "active"

    response = client.post(
        f"/api/admin/events/{event_id}/prizes", headers=ADMIN, json=prize_payload()
    )
    assert response.status_code == 201, response.text
    prize_id = response.json()["id"]
    assert response.json()["real_value"] == 19900

    edited = {**prize_payload(), "name": "升级保温杯", "stock": 8}
    response = client.put(f"/api/admin/prizes/{prize_id}", headers=ADMIN, json=edited)
    assert response.status_code == 200
    assert response.json()["name"] == "升级保温杯"
    assert response.json()["stock"] == 8

    assert client.delete(f"/api/admin/prizes/{prize_id}", headers=ADMIN).status_code == 204
    assert client.get(f"/api/admin/prizes/{prize_id}", headers=ADMIN).status_code == 404


def test_prize_rejects_extra_fields_and_non_https_image() -> None:
    event_id = create_event()
    bad = {**prize_payload(), "category": "未要求字段"}
    assert client.post(f"/api/admin/events/{event_id}/prizes", headers=ADMIN, json=bad).status_code == 422
    bad = {**prize_payload(), "image": "http://example.com/cup.jpg"}
    assert client.post(f"/api/admin/events/{event_id}/prizes", headers=ADMIN, json=bad).status_code == 422
    bad = {**prize_payload(), "jd_url": "http://item.jd.com/100.html"}
    assert client.post(f"/api/admin/events/{event_id}/prizes", headers=ADMIN, json=bad).status_code == 422


def test_prize_tag_roundtrip_and_normalization() -> None:
    event_id = create_event()
    created = client.post(
        f"/api/admin/events/{event_id}/prizes",
        headers=ADMIN,
        json={**prize_payload(), "tag": " 1-数码 "},
    )
    assert created.status_code == 201, created.text
    assert created.json()["tag"] == "1-数码"

    blank = client.post(
        f"/api/admin/events/{event_id}/prizes",
        headers=ADMIN,
        json={**prize_payload(), "name": "无标签奖品", "tag": "   "},
    )
    assert blank.status_code == 201, blank.text
    assert blank.json()["tag"] is None


def test_image_upload_checks_file_content() -> None:
    image_data = io.BytesIO()
    Image.new("RGB", (4, 4), "red").save(image_data, format="PNG")
    response = client.post(
        "/api/admin/uploads/prize-image",
        headers=ADMIN,
        files={"file": ("misleading.jpg", image_data.getvalue(), "image/jpeg")},
    )
    assert response.status_code == 201, response.text
    assert response.json()["image"].startswith("/uploads/prizes/")
    assert response.json()["image"].endswith(".png")

    response = client.post(
        "/api/admin/uploads/prize-image",
        headers=ADMIN,
        files={"file": ("fake.png", b"not an image", "image/png")},
    )
    assert response.status_code == 422


def csv_file(rows: list[list[object]]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(["name", "image", "real_value", "redeem_value", "stock", "description"])
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8")


def test_csv_import_is_atomic_and_export_has_money_format() -> None:
    event_id = create_event()
    valid = csv_file(
        [
            ["保温杯", "https://example.com/cup.jpg", "199.00", 150, 20, "黑色"],
            ["背包", "https://example.com/bag.jpg", "299.90", 250, 10, "蓝色"],
        ]
    )
    response = client.post(
        f"/api/admin/events/{event_id}/prizes/import/validate",
        headers=ADMIN,
        files={"file": ("prizes.csv", valid, "text/csv")},
    )
    assert response.status_code == 200
    assert response.json()["valid"] is True
    assert response.json()["rows"][0]["real_value"] == "199.00"

    response = client.post(
        f"/api/admin/events/{event_id}/prizes/import/confirm",
        headers=ADMIN,
        files={"file": ("prizes.csv", valid, "text/csv")},
    )
    assert response.status_code == 201
    assert response.json() == {"imported": 2}

    export = client.get(f"/api/admin/events/{event_id}/prizes/export?format=csv", headers=ADMIN)
    assert export.status_code == 200
    text = export.content.decode("utf-8-sig")
    assert "199.00" in text and "299.90" in text

    invalid = csv_file(
        [
            ["有效奖品", "https://example.com/ok.jpg", "10.00", 10, 1, ""],
            ["错误奖品", "http://example.com/no.jpg", "10.00", 0, -1, ""],
        ]
    )
    response = client.post(
        f"/api/admin/events/{event_id}/prizes/import/confirm",
        headers=ADMIN,
        files={"file": ("prizes.csv", invalid, "text/csv")},
    )
    assert response.status_code == 422
    with Session(test_engine) as session:
        assert session.scalar(select(func.count(Prize.id)).where(Prize.event_id == event_id)) == 2


def test_prize_is_active_and_admin_list_orders_by_tag() -> None:
    event_id = create_event()
    for name, tag in (("生活", "2-生活"), ("数码", "1-数码"), ("默认", None)):
        response = client.post(
            f"/api/admin/events/{event_id}/prizes",
            headers=ADMIN,
            json={**prize_payload(), "name": name, "tag": tag},
        )
        assert response.status_code == 201, response.text
        assert response.json()["is_active"] is True

    listed = client.get(f"/api/admin/events/{event_id}/prizes", headers=ADMIN).json()
    assert [(prize["tag"], prize["is_active"]) for prize in listed] == [
        ("1-数码", True),
        ("2-生活", True),
        (None, True),
    ]

    # Off-shelf prizes stay visible to admins but are hidden from the public page.
    off_shelf = {**prize_payload(), "name": "下架奖品", "is_active": False}
    assert client.post(f"/api/admin/events/{event_id}/prizes", headers=ADMIN, json=off_shelf).status_code == 201
    listed = client.get(f"/api/admin/events/{event_id}/prizes", headers=ADMIN).json()
    assert listed[-1]["is_active"] is False


def test_prize_batch_tag_and_stock_and_delete() -> None:
    event_id = create_event()
    ids = []
    for name in ("甲", "乙", "丙"):
        response = client.post(
            f"/api/admin/events/{event_id}/prizes",
            headers=ADMIN,
            json={**prize_payload(), "name": name, "stock": 10},
        )
        assert response.status_code == 201, response.text
        ids.append(response.json()["id"])

    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-tag",
        headers=ADMIN,
        json={"ids": ids[:2], "tag": " 1-数码 "},
    ).json() == {"updated": 2}
    listed = client.get(f"/api/admin/events/{event_id}/prizes", headers=ADMIN).json()
    by_id = {prize["id"]: prize for prize in listed}
    assert [by_id[prize_id]["tag"] for prize_id in ids] == ["1-数码", "1-数码", None]

    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-tag",
        headers=ADMIN,
        json={"ids": ids[:2], "tag": None},
    ).json() == {"updated": 2}
    listed = client.get(f"/api/admin/events/{event_id}/prizes", headers=ADMIN).json()
    assert all(prize["tag"] is None for prize in listed)

    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-tag",
        headers=ADMIN,
        json={"ids": [], "tag": "x"},
    ).status_code == 422

    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-stock",
        headers=ADMIN,
        json={"ids": ids, "mode": "delta", "value": 5},
    ).json() == {"updated": 3}
    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-stock",
        headers=ADMIN,
        json={"ids": ids[:2], "mode": "delta", "value": -3},
    ).json() == {"updated": 2}
    listed = client.get(f"/api/admin/events/{event_id}/prizes", headers=ADMIN).json()
    by_id = {prize["id"]: prize for prize in listed}
    assert [by_id[prize_id]["stock"] for prize_id in ids] == [12, 12, 15]

    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-stock",
        headers=ADMIN,
        json={"ids": ids[:1], "mode": "set", "value": 2},
    ).json() == {"updated": 1}
    # The max itself is in range; stepping past it in either direction must fail.
    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-stock",
        headers=ADMIN,
        json={"ids": ids[:1], "mode": "set", "value": 9_223_372_036_854_775_807},
    ).json() == {"updated": 1}
    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-stock",
        headers=ADMIN,
        json={"ids": ids[:1], "mode": "delta", "value": 1},
    ).status_code == 422
    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-stock",
        headers=ADMIN,
        json={"ids": ids[:1], "mode": "set", "value": -9_223_372_036_854_775_809},
    ).status_code == 422

    assert client.post(
        f"/api/admin/events/{event_id}/prizes/batch-delete",
        headers=ADMIN,
        json={"ids": ids[:2]},
    ).json() == {"deleted": 2, "skipped": []}
    with Session(test_engine) as session:
        assert session.scalar(select(func.count(Prize.id)).where(Prize.event_id == event_id)) == 1


def test_csv_import_with_tag_column() -> None:
    event_id = create_event()
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(
        ["name", "image", "real_value", "purchase_value", "redeem_value", "stock", "description", "jd_url", "tag"]
    )
    writer.writerow(["保温杯", "https://example.com/cup.jpg", "199.00", "199.00", 150, 20, "黑色", "", "1-数码"])
    writer.writerow(["背包", "https://example.com/bag.jpg", "299.90", "299.90", 250, 10, "蓝色", "", ""])
    content = stream.getvalue().encode("utf-8")
    response = client.post(
        f"/api/admin/events/{event_id}/prizes/import/confirm",
        headers=ADMIN,
        files={"file": ("prizes.csv", content, "text/csv")},
    )
    assert response.status_code == 201, response.text
    prizes = client.get(f"/api/admin/events/{event_id}/prizes", headers=ADMIN).json()
    assert [prize["tag"] for prize in prizes] == ["1-数码", None]


def test_xlsx_import_and_export_match_csv() -> None:
    event_id = create_event()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "image", "real_value", "redeem_value", "stock", "description"])
    sheet.append(["键盘", "https://example.com/keyboard.jpg", "399.00", 300, 4, "机械键盘"])
    content = io.BytesIO()
    workbook.save(content)
    response = client.post(
        f"/api/admin/events/{event_id}/prizes/import/confirm",
        headers=ADMIN,
        files={"file": ("prizes.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 201, response.text
    exported = client.get(f"/api/admin/events/{event_id}/prizes/export?format=xlsx", headers=ADMIN)
    assert exported.status_code == 200
    output = load_workbook(io.BytesIO(exported.content), read_only=True)
    values = list(output.active.values)
    assert values[1][0] == "键盘"
    assert values[1][2] == "399.00"
    output.close()
