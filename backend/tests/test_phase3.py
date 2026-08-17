import csv
import io
import re

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.main import app
from app.models import NotificationChannel, NotificationJob, RedemptionCode, Winner
from tests.conftest import test_engine
from tests.test_phase2 import create_event


client = TestClient(app)
ADMIN = {"X-Admin-Password": "prizepass-dev-admin"}


def winner_csv(rows: list[list[object]], external: bool = True) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(["external_id", "name", "email", "quota"] if external else ["name", "email", "quota"])
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8")


def post_file(event_id: int, action: str, name: str, content: bytes):
    return client.post(
        f"/api/admin/events/{event_id}/winners/import/{action}",
        headers=ADMIN,
        files={"file": (name, content, "application/octet-stream")},
    )


def test_invalid_rows_reject_entire_winner_import() -> None:
    event_id = create_event()
    content = winner_csv(
        [
            ["A001", "张三", "same@example.com", 100],
            ["A002", "李四", "SAME@example.com", 300],
            ["A003", "王五", "invalid-email", "1.5"],
        ]
    )
    preview = post_file(event_id, "validate", "winners.csv", content)
    assert preview.status_code == 200
    assert preview.json()["valid"] is False
    assert {error["field"] for error in preview.json()["errors"]} >= {"email", "quota"}
    confirmed = post_file(event_id, "confirm", "winners.csv", content)
    assert confirmed.status_code == 422
    with Session(test_engine) as session:
        assert session.scalar(select(func.count(Winner.id))) == 0
        assert session.scalar(select(func.count(RedemptionCode.id))) == 0
        assert session.scalar(select(func.count(NotificationJob.id))) == 0


def test_import_generates_codes_and_same_text_channel_jobs() -> None:
    event_id = create_event()
    content = winner_csv(
        [
            ["A001", "张三", "zhangsan@example.com", 100],
            ["A002", "李四", "lisi@example.com", 300],
            ["A003", "王五", "wangwu@example.com", 500],
        ]
    )
    preview = post_file(event_id, "validate", "winners.csv", content)
    assert preview.status_code == 200
    assert preview.json()["valid"] is True
    assert preview.json()["count"] == 3
    assert preview.json()["quota_total"] == 900
    confirmed = post_file(event_id, "confirm", "winners.csv", content)
    assert confirmed.status_code == 201, confirmed.text
    assert confirmed.json() == {"imported": 3}

    with Session(test_engine) as session:
        winners = session.scalars(select(Winner).order_by(Winner.id)).all()
        codes = session.scalars(select(RedemptionCode).order_by(RedemptionCode.id)).all()
        jobs = session.scalars(select(NotificationJob).order_by(NotificationJob.winner_id, NotificationJob.id)).all()
        assert [winner.quota for winner in winners] == [100, 300, 500]
        assert len({code.code for code in codes}) == 3
        assert all(re.fullmatch(r"[ABCDEFGHJKLMNPQRSTUVWXYZ23456789]{12}", code.code) for code in codes)
        assert len(jobs) == 6
        for winner in winners:
            pair = [job for job in jobs if job.winner_id == winner.id]
            assert {job.channel for job in pair} == {NotificationChannel.EMAIL, NotificationChannel.WEBHOOK}
            assert pair[0].text_rendered == pair[1].text_rendered

    listed = client.get(f"/api/admin/events/{event_id}/winners", headers=ADMIN)
    assert listed.status_code == 200
    assert len(listed.json()) == 3
    assert listed.json()[0]["email_notification_status"] == "pending"
    assert listed.json()[0]["webhook_notification_status"] == "pending"


def test_duplicate_database_identity_does_not_create_more_jobs() -> None:
    event_id = create_event()
    first = winner_csv([["A001", "张三", "first@example.com", 100]])
    assert post_file(event_id, "confirm", "winners.csv", first).status_code == 201
    duplicate_external = winner_csv([["A001", "另一人", "other@example.com", 200]])
    response = post_file(event_id, "confirm", "winners.csv", duplicate_external)
    assert response.status_code == 422
    with Session(test_engine) as session:
        assert session.scalar(select(func.count(Winner.id))) == 1
        assert session.scalar(select(func.count(NotificationJob.id))) == 2


def test_xlsx_without_external_id_and_both_exports() -> None:
    event_id = create_event()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "email", "quota"])
    sheet.append(["  陈晨  ", "CHEN@example.com ", 250])
    source = io.BytesIO()
    workbook.save(source)
    response = post_file(event_id, "confirm", "winners.xlsx", source.getvalue())
    assert response.status_code == 201, response.text

    csv_export = client.get(f"/api/admin/events/{event_id}/winners/export?format=csv", headers=ADMIN)
    assert csv_export.status_code == 200
    assert "chen@example.com" in csv_export.content.decode("utf-8-sig")
    xlsx_export = client.get(f"/api/admin/events/{event_id}/winners/export?format=xlsx", headers=ADMIN)
    assert xlsx_export.status_code == 200
    output = load_workbook(io.BytesIO(xlsx_export.content), read_only=True)
    values = list(output.active.values)
    assert values[1][1:5] == ("陈晨", "chen@example.com", None, 250)
    output.close()
